import re
import openpyxl
file_path = 'C:\\Users\\86139\\Desktop\\hpl.txt'  # 将路径替换为你的文件路径
with open(file_path, 'r', encoding='utf-8') as file:
    file_contents = file.read()
file_contents = re.sub(r'[ \t]+', '', file_contents)
lines = file_contents.splitlines()

# title = []
# for line in lines:
#     if not line.startswith("2"):
#         title.append(line)
#
# l=0
# candj=[]
# while l<title.__len__():
#     if title[l][1]=='c' or title[l][1]=='j':
#         candj.append(title[l+1])
#     l=l+2
# print(candj.__len__())
#
# l=0
# listi=[]
# while l<title.__len__():
#     if title[l][1]=='i':
#         listi.append(title[l+1])
#     l=l+2
#
#
# print(listi)
#
#
# list=[]
# list2=[]
# for i in range(candj.__len__()):
#     dot_index1 = candj[i].find(".")
#     substring1 = candj[i][:dot_index1 + 1]
#     p=0
#     for j in range(listi.__len__()):
#         dot_index2 = listi[j].find(".")
#         substring2 = listi[j][:dot_index2 + 1]
#         if substring1==substring2:
#             list.append(listi[j][listi[j].find("abs") :listi[j].find("(")])
#             list2.append("https://arxiv.org/pdf/"+listi[j][listi[j].find("/")+1:listi[j].find("(")])
#             p=1
#     if p==0:
#         list.append("")
#         list2.append("")
#     else:
#         p=0
# print(list)
# print(list2)

candj = []
for line in lines:
    if line.startswith("[") and line[1] in ["c", "j"]:
        start_index = line.find("]") + 1  # 找到"]"后的下一个位置
        end_index = line.find(":")  # 找到":"的位置
        content = line[start_index:end_index]  # 使用切片提取内容
        result = ""
        for i, char in enumerate(content):
            if char.isupper():
                if i > 0 and content[i - 1] != ",":
                    result += " "
                result += char
            elif char==",":
                result += char
                result += " "
            else:
                result += char
        candj.append(result)

print(candj)
wb = openpyxl.Workbook()
ws = wb.active

for index, value in enumerate(candj, start=1):
    cell = ws.cell(row=index, column=1)
    cell.value = value
# 保存 Excel 文档
wb.save("output3.xlsx")
# wb = openpyxl.Workbook()
# ws = wb.active

# for index, value in enumerate(list, start=1):
#     cell = ws.cell(row=index, column=1)
#     cell.value = value
# for index, value in enumerate(list2, start=1):
#     cell = ws.cell(row=index, column=2)
#     cell.value = value
# # 保存 Excel 文档
# wb.save("output2.xlsx")






